import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


def loginopencart(username, password):
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT, "My Account").click()
    driver.find_element(By.LINK_TEXT, "Login").click()
    time.sleep(5)
    print(driver.title)
    if driver.title == "Your store":
        print("matched")
    else:
        print("not matched")
    driver.find_element(By.ID, "input-email").send_keys(username)
    driver.find_element(By.ID, "input-password").send_keys(password)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(5)
    driver.quit()
excelpath="C:\\Users\\Administrator\\Documents\\Book1.xlsx"
workbook=openpyxl.load_workbook(excelpath)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
for i in range(2, rows):
    uname = sheet.cell(row=1, column=1).value
    pword = sheet.cell(row=1, column=1).value
loginopencart(uname, pword)


