
import time

import openpyxl

from selenium import webdriver

from webdriver_manager.chrome import ChromeDriverManager

import pytest

from pageobjects import loginpage

@pytest.fixture()

def setup():

    print("run before every method")

def test_def1(setup):

    print("hello world")

def test_loginopencart(setup):

    excelpath = "C:\\Users\\Administrator\\Documents\\Book1.xlsx"

    workbook = openpyxl.load_workbook(excelpath)

    sheet = workbook.active

    rows = sheet.max_row

    cols = sheet.max_column

    print(rows)

    for i in range(2, rows + 1):

        print(i)

        uname = sheet.cell(row=i, column=1).value

        pword = sheet.cell(row=i, column=2).value

        driver = webdriver.Chrome(ChromeDriverManager().install())

        driver.get("https://demo.opencart.com/")

        lp = loginpage(driver)

        # driver.find_element(By.LINK_TEXT,"My Account").click()

        lp.clickonmyaccount()

        # print(driver.find_element(By.LINK_TEXT,"My Account()
        # print(driver.find_element(By.LINK_TEXT,"My Account").get_attribute("text"))

        # driver.find_element(By.LINK_TEXT,"Login").click()

        lp.clickonloginlink()

        time.sleep(5)

        lp.enteremail(uname)

        # driver.find_element(By.ID,"input-password").send_keys(password)

        lp.enterpassword(pword)

        # driver.find_element(By.XPATH,"//button[@type='submit']").click()

        lp.clickonsubmit()

        time.sleep(5)